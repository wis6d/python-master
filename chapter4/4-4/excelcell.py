#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Sat Feb 22 15:38:51 2020

@author: hideyuki.suzuki
"""

from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws.title = 'Sample'
ws['A1'].value = 'Hello Excel'

for i in range(2, 10):
    ws.cell(row=i, column=1).value = 'No,' + str(i - 1)
wb.save('data.xlsx')
print('saved.')