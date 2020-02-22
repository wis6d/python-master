#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Sat Feb 22 15:48:17 2020

@author: hideyuki.suzuki
"""

import random
from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws.title = 'Sample'

def makeData():
    ws['A1'].value = '支店'
    ws['B1'].value = '売上'
    
    data =  ['東京', '大阪', '名古屋', '札幌', '仙台']
    for i in range(2,7):
        ws.cell(row=i, column=1).value = data[i - 2]
        _cell = ws.cell(row=i, column=2)
        _cell.number_format = '#,##0'
        _cell.value = random.randint(1, 100) * 10
    ws['A10'].value = '合計'
    ws['B10'].number_format = '#,##0'
    ws['B10'].value = '=SUM(B2:B6)'

makeData()

wb.save('data.xlsx')
print('saved.')
    